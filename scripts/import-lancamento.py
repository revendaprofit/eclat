# -*- coding: utf-8 -*-
"""
Importador do lançamento — lê docs/lancamento-familia-blackout.xlsx e cadastra
tudo no Medusa (produção): coleção, categorias, produtos com variantes Cor×Tamanho,
preços BRL, estoque no CD Brasil e foto inicial (render aprovado). Também
despublica os 4 produtos-exemplo do seed.

Requisitos: pip install openpyxl requests
Credenciais: apps/cockpit/.env.local -> MEDUSA_ADMIN_EMAIL / MEDUSA_ADMIN_PASSWORD
             (a senha de PRODUÇÃO — colocar manualmente, nunca commitar)
Uso:        python scripts/import-lancamento.py [--dry-run]
"""
import io, os, sys, unicodedata

try:
    import requests
    from openpyxl import load_workbook
except ImportError:
    print("Rode antes: pip install openpyxl requests")
    sys.exit(1)

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(RAIZ, "docs", "lancamento-familia-blackout.xlsx")
RENDERS = os.path.join(RAIZ, "docs", "design", "familia-blackout-verde-licor")
BASE = os.environ.get("MEDUSA_URL", "https://endearing-enthusiasm-production-775b.up.railway.app")
COLECAO = ("Família Blackout — Verde Exército e Licor", "familia-blackout")
SEEDS_DESPUBLICAR = ["top-aurora", "short-solene", "conjunto-luz", "legging-resplendor"]
CATEGORIAS = {"top": "Tops", "shorts": "Shorts", "legging": "Leggings", "macaquinho": "Macaquinhos"}
TAMANHOS = ["P", "M", "G", "GG"]
DRY = "--dry-run" in sys.argv

def slugify(s):
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn").lower()
    return "-".join("".join(c if c.isalnum() else " " for c in s).split())

def env_cockpit():
    env = {}
    p = os.path.join(RAIZ, "apps", "cockpit", ".env.local")
    for line in io.open(p, encoding="utf-8").read().splitlines():
        if "=" in line and not line.startswith("#"):
            k, _, v = line.partition("=")
            env[k.strip()] = v.strip()
    return env

def main():
    env = env_cockpit()
    email, senha = env.get("MEDUSA_ADMIN_EMAIL"), env.get("MEDUSA_ADMIN_PASSWORD")
    assert email and senha, "MEDUSA_ADMIN_EMAIL/PASSWORD ausentes no apps/cockpit/.env.local"

    r = requests.post(BASE + "/auth/user/emailpass", json={"email": email, "password": senha}, timeout=30)
    assert r.ok and r.json().get("token"), "login admin falhou (%s): confira a senha de PRODUÇÃO no .env.local" % r.status_code
    H = {"Authorization": "Bearer " + r.json()["token"]}
    print("login admin OK:", email)

    # planilha -> agrupa por modelo
    ws = load_workbook(XLSX)["Produtos"]
    cols = [c.value for c in ws[1]]
    modelos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(cols, row))
        if not d.get("MODELO") or not d.get("COR"):
            continue
        preco = float(d.get("PRECO_REAIS") or 0)
        assert preco > 0, "PRECO_REAIS vazio para %s / %s — preencha a planilha" % (d["MODELO"], d["COR"])
        modelos.setdefault(d["MODELO"], []).append(d)

    print("modelos na planilha:", len(modelos), "| combinações cor:", sum(len(v) for v in modelos.values()))
    if DRY:
        for m, rows in modelos.items():
            print("-", m, "->", ", ".join("%s R$%.2f" % (x["COR"], float(x["PRECO_REAIS"])) for x in rows))
        return

    # stock location
    r = requests.get(BASE + "/admin/stock-locations", headers=H, timeout=30)
    sloc = r.json()["stock_locations"][0]["id"]
    print("stock location:", sloc)

    # coleção
    r = requests.get(BASE + "/admin/collections?handle=" + COLECAO[1], headers=H, timeout=30)
    cols_found = r.json().get("collections", [])
    if cols_found:
        col_id = cols_found[0]["id"]
    else:
        r = requests.post(BASE + "/admin/collections", headers=H, json={"title": COLECAO[0], "handle": COLECAO[1]}, timeout=30)
        col_id = r.json()["collection"]["id"]
    print("coleção:", col_id)

    # categorias (cria se faltar)
    r = requests.get(BASE + "/admin/product-categories?limit=100", headers=H, timeout=30)
    existentes = {c["name"].lower(): c["id"] for c in r.json().get("product_categories", [])}
    cat_ids = {}
    for chave, nome in CATEGORIAS.items():
        if nome.lower() in existentes:
            cat_ids[chave] = existentes[nome.lower()]
        else:
            r = requests.post(BASE + "/admin/product-categories", headers=H, json={"name": nome, "is_active": True}, timeout=30)
            cat_ids[chave] = r.json()["product_category"]["id"]
    print("categorias:", cat_ids)

    # handles existentes (não duplicar)
    r = requests.get(BASE + "/admin/products?limit=200&fields=id,handle", headers=H, timeout=30)
    handles = {p["handle"]: p["id"] for p in r.json().get("products", [])}

    RENDER_MAP = {
        "top-radiance": "1-top-radiance.png", "top-lumina": "2-top-lumina.png",
        "top-prisma": "3-top-prisma.png", "shorts-radiance": "4-shorts-radiance.png",
        "shorts-lumina": "5-shorts-lumina.png", "shorts-prisma": "6-shorts-prisma.png",
        "legging-vertice": "7-legging-vertice.png", "macaquinho-prisma": "8-macaquinho-prisma.png",
    }

    for modelo, rows in modelos.items():
        handle = slugify(modelo)
        if handle in handles:
            print("PULANDO (já existe):", handle)
            continue
        cores = [x["COR"] for x in rows]
        cat = rows[0]["CATEGORIA"]

        # foto inicial: render aprovado
        image_url = None
        arq = RENDER_MAP.get(handle)
        if arq and os.path.exists(os.path.join(RENDERS, arq)):
            with open(os.path.join(RENDERS, arq), "rb") as f:
                up = requests.post(BASE + "/admin/uploads", headers=H,
                                   files=[("files", (arq, f, "image/png"))], timeout=120)
            if up.ok:
                image_url = up.json()["files"][0]["url"]

        variants = []
        for x in rows:
            preco = float(x["PRECO_REAIS"])  # Medusa v2: reais decimais, NAO centavos
            for t in TAMANHOS:
                variants.append({
                    "title": "%s / %s" % (t, x["COR"]),
                    "sku": "%s-%s" % (x["SKU"], t),
                    "options": {"Tamanho": t, "Cor": x["COR"]},
                    "manage_inventory": True,
                    "prices": [{"amount": preco, "currency_code": "brl"}],
                })

        payload = {
            "title": modelo,
            "handle": handle,
            "status": "published" if str(rows[0].get("PUBLICAR", "sim")).lower().startswith("s") else "draft",
            "description": rows[0].get("DESCRICAO") or modelo,
            "options": [
                {"title": "Tamanho", "values": TAMANHOS},
                {"title": "Cor", "values": cores},
            ],
            "variants": variants,
            "collection_id": col_id,
            "categories": [{"id": cat_ids[cat]}],
        }
        if image_url:
            payload["images"] = [{"url": image_url}]
            payload["thumbnail"] = image_url

        r = requests.post(BASE + "/admin/products", headers=H, json=payload, timeout=120)
        assert r.ok, "criar %s falhou: %s %s" % (modelo, r.status_code, r.text[:300])
        prod = r.json()["product"]
        print("CRIADO:", modelo, prod["id"], "| variantes:", len(prod.get("variants", [])))

        # estoque por variante (mapa SKU -> quantidade do tamanho)
        est = {}
        for x in rows:
            for t in TAMANHOS:
                est["%s-%s" % (x["SKU"], t)] = int(x.get("ESTOQUE_" + t) or 0)
        r = requests.get(BASE + "/admin/products/" + prod["id"] +
                         "?fields=id,*variants.sku,*variants.inventory_items", headers=H, timeout=60)
        for v in r.json()["product"]["variants"]:
            qtd = est.get(v.get("sku"), 0)
            for ii in (v.get("inventory_items") or []):
                iid = ii.get("inventory_item_id") or ii.get("inventory", {}).get("id")
                if iid:
                    requests.post(BASE + "/admin/inventory-items/%s/location-levels" % iid,
                                  headers=H, json={"location_id": sloc, "stocked_quantity": qtd}, timeout=60)
        print("  estoque lançado no CD Brasil")

    # despublica seeds
    for h in SEEDS_DESPUBLICAR:
        pid = handles.get(h)
        if pid:
            requests.post(BASE + "/admin/products/" + pid, headers=H, json={"status": "draft"}, timeout=60)
            print("DESPUBLICADO seed:", h)

    print()
    print("IMPORT CONCLUIDO. Rode scripts/fix-import.py se precisar reprocessar precos/estoque/fotos.")

if __name__ == "__main__":
    main()
