# -*- coding: utf-8 -*-
"""
Correção do import da Família Blackout:
1) preços em REAIS decimais (Medusa v2) — estava x100
2) estoque por SKU via inventory-items (upsert de location-level)
3) fotos re-hospedadas no Supabase Storage (público/permanente) e religadas
Uso: python scripts/fix-import.py
"""
import io, os, requests
from openpyxl import load_workbook

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.environ.get("MEDUSA_URL", "https://endearing-enthusiasm-production-775b.up.railway.app")
XLSX = os.path.join(RAIZ, "docs", "lancamento-familia-blackout.xlsx")
RENDERS = os.path.join(RAIZ, "docs", "design", "familia-blackout-verde-licor")
TAMANHOS = ["P", "M", "G", "GG"]
HANDLES = ["top-radiance", "top-lumina", "top-prisma", "shorts-radiance",
           "shorts-lumina", "shorts-prisma", "legging-vertice", "macaquinho-prisma"]
RENDER_MAP = {
    "top-radiance": "1-top-radiance.png", "top-lumina": "2-top-lumina.png",
    "top-prisma": "3-top-prisma.png", "shorts-radiance": "4-shorts-radiance.png",
    "shorts-lumina": "5-shorts-lumina.png", "shorts-prisma": "6-shorts-prisma.png",
    "legging-vertice": "7-legging-vertice.png", "macaquinho-prisma": "8-macaquinho-prisma.png",
}

def env_cockpit():
    env = {}
    for line in io.open(os.path.join(RAIZ, "apps", "cockpit", ".env.local"), encoding="utf-8").read().splitlines():
        if "=" in line and not line.startswith("#"):
            k, _, v = line.partition("=")
            env[k.strip()] = v.strip()
    return env

def main():
    env = env_cockpit()
    r = requests.post(BASE + "/auth/user/emailpass",
                      json={"email": env["MEDUSA_ADMIN_EMAIL"], "password": env["MEDUSA_ADMIN_PASSWORD"]}, timeout=30)
    H = {"Authorization": "Bearer " + r.json()["token"]}
    SB_URL, SB_KEY = env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]
    SBH = {"apikey": SB_KEY, "Authorization": "Bearer " + SB_KEY}

    # planilha: preço por SKU-base e estoque por SKU completo
    ws = load_workbook(XLSX)["Produtos"]
    cols = [c.value for c in ws[1]]
    preco_por_sku, estoque_por_sku = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(cols, row))
        if not d.get("SKU"):
            continue
        for t in TAMANHOS:
            sku = "%s-%s" % (d["SKU"], t)
            preco_por_sku[sku] = float(d["PRECO_REAIS"])
            estoque_por_sku[sku] = int(d.get("ESTOQUE_" + t) or 0)

    # stock location
    sloc = requests.get(BASE + "/admin/stock-locations", headers=H, timeout=30).json()["stock_locations"][0]["id"]

    for handle in HANDLES:
        r = requests.get(BASE + "/admin/products?handle=" + handle + "&fields=id,title,thumbnail,*variants.id,*variants.sku",
                         headers=H, timeout=60)
        prods = r.json().get("products", [])
        if not prods:
            print("NAO ACHOU", handle)
            continue
        p = prods[0]
        print("==", p["title"])

        # 1) PRECOS em reais decimais
        ok_precos = 0
        for v in p["variants"]:
            preco = preco_por_sku.get(v.get("sku"))
            if preco is None:
                continue
            rr = requests.post(BASE + "/admin/products/%s/variants/%s" % (p["id"], v["id"]),
                               headers=H, json={"prices": [{"amount": preco, "currency_code": "brl"}]}, timeout=60)
            ok_precos += 1 if rr.ok else 0
        print("   precos corrigidos:", ok_precos, "/", len(p["variants"]))

        # 2) ESTOQUE via inventory item por SKU
        ok_est = 0
        for v in p["variants"]:
            sku = v.get("sku")
            qtd = estoque_por_sku.get(sku, 0)
            ri = requests.get(BASE + "/admin/inventory-items?q=" + sku + "&limit=5", headers=H, timeout=60)
            items = [i for i in ri.json().get("inventory_items", []) if i.get("sku") == sku]
            if not items:
                print("   SEM inventory item p/", sku)
                continue
            iid = items[0]["id"]
            # upsert do nivel
            rl = requests.get(BASE + "/admin/inventory-items/%s/location-levels" % iid, headers=H, timeout=60)
            niveis = rl.json().get("inventory_levels", []) if rl.ok else []
            existe = any(n.get("location_id") == sloc for n in niveis)
            if existe:
                ru = requests.post(BASE + "/admin/inventory-items/%s/location-levels/%s" % (iid, sloc),
                                   headers=H, json={"stocked_quantity": qtd}, timeout=60)
            else:
                ru = requests.post(BASE + "/admin/inventory-items/%s/location-levels" % iid,
                                   headers=H, json={"location_id": sloc, "stocked_quantity": qtd}, timeout=60)
            ok_est += 1 if ru.ok else 0
        print("   estoque lancado:", ok_est, "/", len(p["variants"]))

        # 3) FOTO no Supabase Storage
        arq = RENDER_MAP.get(handle)
        caminho = os.path.join(RENDERS, arq) if arq else None
        if caminho and os.path.exists(caminho):
            dest = "products/%s.png" % handle
            with open(caminho, "rb") as f:
                up = requests.post(SB_URL + "/storage/v1/object/site/" + dest,
                                   headers=dict(SBH, **{"Content-Type": "image/png", "x-upsert": "true"}),
                                   data=f.read(), timeout=120)
            if up.ok:
                url = SB_URL + "/storage/v1/object/public/site/" + dest
                rr = requests.post(BASE + "/admin/products/" + p["id"], headers=H,
                                   json={"images": [{"url": url}], "thumbnail": url}, timeout=60)
                print("   foto:", "OK" if rr.ok else "FALHOU %s" % rr.status_code)
            else:
                print("   upload supabase falhou:", up.status_code, up.text[:120])

    print()
    print("CORRECAO CONCLUIDA")

if __name__ == "__main__":
    main()
